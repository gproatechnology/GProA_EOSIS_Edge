from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
import json
import io
import asyncio
from pathlib import Path
from pydantic import BaseModel, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from edge_rules import EDGE_WBS, get_rule, get_all_rules, validate_project_wbs, get_project_coverage
from edge_processors import run_specialized_processor
from pdf_generator import generate_edge_report

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ── DATABASE: SQLite (default) or MongoDB if MONGO_URL provided ─────
MONGO_URL = os.getenv('MONGO_URL')
DB_NAME = os.getenv('DB_NAME', 'gproa_edge')

if MONGO_URL:
    # MongoDB mode (requires MONGO_URL)
    from motor.motor_asyncio import AsyncIOMotorClient
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    logger = logging.getLogger(__name__)
    logger.info("Using MongoDB database")
else:
    # SQLite mode (default for demo)
    import aiosqlite
    db = None  # Not used in SQLite mode
    DATA_DIR = ROOT_DIR / "data"
    DATA_DIR.mkdir(exist_ok=True)
    DB_PATH = DATA_DIR / f"{DB_NAME}.db"
    logger = logging.getLogger(__name__)
    logger.info(f"Using SQLite database at {DB_PATH}")

# ── OPENAI: real client or mock/demo mode ─────────────────────────────
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY') or os.getenv('EMERGENT_LLM_KEY')
DEMO_MODE = os.getenv('DEMO_MODE', 'false').lower() == 'true'

if OPENAI_API_KEY and not DEMO_MODE:
    from openai import AsyncOpenAI
    openai_client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    logger.info("Using OpenAI API")
else:
    openai_client = None
    logger.info("Demo mode: using mock AI responses")

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# In-memory job tracker for batch progress
processing_jobs = {}

class UnifiedDB:
    """Wrapper that provides same interface for MongoDB and SQLite."""

    def __init__(self):
        self.mode = "mongodb" if MONGO_URL else "sqlite"
        self.conn = None  # Lazy-init on first use

    async def _ensure_sqlite(self):
        """Initialize SQLite connection and schema on first use."""
        if self.conn is None:
            print(">>> Initializing SQLite connection...")
            self.conn = await aiosqlite.connect(DB_PATH)
            self.conn.row_factory = aiosqlite.Row
            await self.conn.execute("""
                CREATE TABLE IF NOT EXISTS projects (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    typology TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    file_count INTEGER DEFAULT 0,
                    processed_count INTEGER DEFAULT 0
                )
            """)
            await self.conn.execute("""
                CREATE TABLE IF NOT EXISTS files (
                    id TEXT PRIMARY KEY,
                    project_id TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    content_text TEXT,
                    status TEXT DEFAULT 'pending',
                    category_edge TEXT,
                    measure_edge TEXT,
                    doc_type TEXT,
                    confidence REAL,
                    watts REAL,
                    lumens REAL,
                    tipo_equipo TEXT,
                    marca TEXT,
                    modelo TEXT,
                    areas TEXT,
                    specialized_data TEXT,
                    uploaded_at TEXT NOT NULL
                )
            """)
            await self.conn.execute("CREATE INDEX IF NOT EXISTS idx_files_project ON files(project_id)")
            await self.conn.execute("CREATE INDEX IF NOT EXISTS idx_files_status ON files(status)")
            await self.conn.commit()
            print(">>> SQLite database ready")

    async def projects_insert_one(self, doc: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            await db.projects.insert_one(doc)
        else:
            await self.conn.execute("""
                INSERT INTO projects (id, name, typology, created_at, file_count, processed_count)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                doc["id"], doc["name"], doc["typology"], doc["created_at"],
                doc.get("file_count", 0), doc.get("processed_count", 0)
            ))
            await self.conn.commit()

    async def projects_find_one(self, query: dict, projection: dict = None):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            return await db.projects.find_one(query, proj)
        else:
            columns = ["id", "name", "typology", "created_at", "file_count", "processed_count"]
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT {', '.join(columns)} FROM projects WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                return dict(row) if row else None

    async def projects_find(self, query: dict = None, projection: dict = None):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            cursor = db.projects.find(query or {}, proj or {})
            return await cursor.to_list(100)
        else:
            columns = ["id", "name", "typology", "created_at", "file_count", "processed_count"]
            sql = f"SELECT {', '.join(columns)} FROM projects"
            params = []
            if query:
                where = [f"{k}=?" for k in query.keys()]
                params = list(query.values())
                sql += " WHERE " + " AND ".join(where)
            async with self.conn.execute(sql, params) as cur:
                rows = await cur.fetchall()
                return [dict(r) for r in rows]

    async def projects_delete_one(self, query: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            return await db.projects.delete_one(query)
        else:
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM projects WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()
                return type('Result', (), {'deleted_count': cur.rowcount})()

    async def files_insert_one(self, doc: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            await db.files.insert_one(doc)
        else:
            areas = json.dumps(doc.get("areas")) if doc.get("areas") else None
            specialized = json.dumps(doc.get("specialized_data")) if doc.get("specialized_data") else None
            await self.conn.execute("""
                INSERT INTO files (
                    id, project_id, filename, file_size, content_text, status,
                    category_edge, measure_edge, doc_type, confidence,
                    watts, lumens, tipo_equipo, marca, modelo,
                    areas, specialized_data, uploaded_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                doc["id"], doc["project_id"], doc["filename"], doc["file_size"],
                doc.get("content_text"), doc.get("status", "pending"),
                doc.get("category_edge"), doc.get("measure_edge"), doc.get("doc_type"),
                doc.get("confidence"),
                doc.get("watts"), doc.get("lumens"), doc.get("tipo_equipo"),
                doc.get("marca"), doc.get("modelo"),
                areas, specialized, doc.get("uploaded_at")
            ))
            await self.conn.commit()

    async def files_find_one(self, query: dict, projection: dict = None):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            return await db.files.find_one(query, proj)
        else:
            columns = ["id", "project_id", "filename", "file_size", "content_text", "status",
                       "category_edge", "measure_edge", "doc_type", "confidence",
                       "watts", "lumens", "tipo_equipo", "marca", "modelo",
                       "areas", "specialized_data", "uploaded_at"]
            if projection and "content_text" in projection and projection["content_text"] == 0:
                columns = [c for c in columns if c != "content_text"]
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT {', '.join(columns)} FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                if row:
                    return self._deserialize_file_row(dict(row))
            return None

    async def files_find(self, query: dict = None, projection: dict = None):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            cursor = db.files.find(query or {}, proj or {})
            return await cursor.to_list(500)
        else:
            columns = ["id", "project_id", "filename", "file_size", "status",
                       "category_edge", "measure_edge", "doc_type", "confidence",
                       "watts", "lumens", "tipo_equipo", "marca", "modelo",
                       "areas", "specialized_data", "uploaded_at"]
            if projection and "content_text" in projection and projection["content_text"] == 0:
                columns = [c for c in columns if c != "content_text"]
            sql = f"SELECT {', '.join(columns)} FROM files"
            params = []
            if query:
                where = [f"{k}=?" for k in query.keys()]
                params = list(query.values())
                sql += " WHERE " + " AND ".join(where)
            async with self.conn.execute(sql, params) as cur:
                rows = await cur.fetchall()
                return [self._deserialize_file_row(dict(r)) for r in rows]

    def _deserialize_file_row(self, row: dict) -> dict:
        if "areas" in row and row["areas"]:
            try:
                row["areas"] = json.loads(row["areas"])
            except:
                row["areas"] = None
        if "specialized_data" in row and row["specialized_data"]:
            try:
                row["specialized_data"] = json.loads(row["specialized_data"])
            except:
                row["specialized_data"] = None
        return row

    async def files_update_one(self, query: dict, update: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            await db.files.update_one(query, update)
        else:
            set_dict = update.get("$set", {})
            if "areas" in set_dict and isinstance(set_dict["areas"], list):
                set_dict["areas"] = json.dumps(set_dict["areas"])
            if "specialized_data" in set_dict and isinstance(set_dict["specialized_data"], dict):
                set_dict["specialized_data"] = json.dumps(set_dict["specialized_data"])
            set_clause = ", ".join([f"{k}=?" for k in set_dict.keys()])
            params = list(set_dict.values())
            where = [f"{k}=?" for k in query.keys()]
            params.extend(list(query.values()))
            sql = f"UPDATE files SET {set_clause} WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()

    async def files_delete_one(self, query: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            return await db.files.delete_one(query)
        else:
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()
                return type('Result', (), {'deleted_count': cur.rowcount})()

    async def files_delete_many(self, query: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            await db.files.delete_many(query)
        else:
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params):
                await self.conn.commit()

    async def files_count_documents(self, query: dict):
        if self.mode == "sqlite":
            await self._ensure_sqlite()
        if self.mode == "mongodb":
            return await db.files.count_documents(query)
        else:
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT COUNT(*) as cnt FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                return row["cnt"] if row else 0

# Initialize unified DB
udb = UnifiedDB()

class UnifiedDB:
    """Wrapper that provides same interface for MongoDB and SQLite."""

    def __init__(self):
        self.mode = "mongodb" if MONGO_URL else "sqlite"
        self.conn = None  # Lazy-init on first use

    async def _ensure_sqlite(self):
        """Initialize SQLite connection and schema on first use."""
        if self.conn is None:
            print(">>> Initializing SQLite connection...")
            self.conn = await aiosqlite.connect(DB_PATH)
            self.conn.row_factory = aiosqlite.Row
            # Create tables
            await self.conn.execute("""
                CREATE TABLE IF NOT EXISTS projects (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    typology TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    file_count INTEGER DEFAULT 0,
                    processed_count INTEGER DEFAULT 0
                )
            """)
            await self.conn.execute("""
                CREATE TABLE IF NOT EXISTS files (
                    id TEXT PRIMARY KEY,
                    project_id TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    content_text TEXT,
                    status TEXT DEFAULT 'pending',
                    category_edge TEXT,
                    measure_edge TEXT,
                    doc_type TEXT,
                    confidence REAL,
                    watts REAL,
                    lumens REAL,
                    tipo_equipo TEXT,
                    marca TEXT,
                    modelo TEXT,
                    areas TEXT,
                    specialized_data TEXT,
                    uploaded_at TEXT NOT NULL
                )
            """)
            await self.conn.execute("CREATE INDEX IF NOT EXISTS idx_files_project ON files(project_id)")
            await self.conn.execute("CREATE INDEX IF NOT EXISTS idx_files_status ON files(status)")
            await self.conn.commit()
            print(">>> SQLite database ready")

    async def projects_insert_one(self, doc: dict):
        if self.mode == "mongodb":
            await db.projects.insert_one(doc)
        else:
            await self._ensure_sqlite()
            await self.conn.execute("""
                INSERT INTO projects (id, name, typology, created_at, file_count, processed_count)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                doc["id"], doc["name"], doc["typology"], doc["created_at"],
                doc.get("file_count", 0), doc.get("processed_count", 0)
            ))
            await self.conn.commit()

    async def projects_find_one(self, query: dict, projection: dict = None):
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            return await db.projects.find_one(query, proj)
        else:
            await self._ensure_sqlite()
            columns = ["id", "name", "typology", "created_at", "file_count", "processed_count"]
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT {', '.join(columns)} FROM projects WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                return dict(row) if row else None

    async def projects_find(self, query: dict = None, projection: dict = None):
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            cursor = db.projects.find(query or {}, proj or {})
            return await cursor.to_list(100)
        else:
            await self._ensure_sqlite()
            columns = ["id", "name", "typology", "created_at", "file_count", "processed_count"]
            sql = f"SELECT {', '.join(columns)} FROM projects"
            params = []
            if query:
                where = [f"{k}=?" for k in query.keys()]
                params = list(query.values())
                sql += " WHERE " + " AND ".join(where)
            async with self.conn.execute(sql, params) as cur:
                rows = await cur.fetchall()
                return [dict(r) for r in rows]

    async def projects_delete_one(self, query: dict):
        if self.mode == "mongodb":
            return await db.projects.delete_one(query)
        else:
            await self._ensure_sqlite()
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM projects WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()
                return type('Result', (), {'deleted_count': cur.rowcount})()

    async def files_insert_one(self, doc: dict):
        if self.mode == "mongodb":
            await db.files.insert_one(doc)
        else:
            await self._ensure_sqlite()
            areas = json.dumps(doc.get("areas")) if doc.get("areas") else None
            specialized = json.dumps(doc.get("specialized_data")) if doc.get("specialized_data") else None
            await self.conn.execute("""
                INSERT INTO files (
                    id, project_id, filename, file_size, content_text, status,
                    category_edge, measure_edge, doc_type, confidence,
                    watts, lumens, tipo_equipo, marca, modelo,
                    areas, specialized_data, uploaded_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                doc["id"], doc["project_id"], doc["filename"], doc["file_size"],
                doc.get("content_text"), doc.get("status", "pending"),
                doc.get("category_edge"), doc.get("measure_edge"), doc.get("doc_type"),
                doc.get("confidence"),
                doc.get("watts"), doc.get("lumens"), doc.get("tipo_equipo"),
                doc.get("marca"), doc.get("modelo"),
                areas, specialized, doc.get("uploaded_at")
            ))
            await self.conn.commit()

    async def files_find_one(self, query: dict, projection: dict = None):
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            return await db.files.find_one(query, proj)
        else:
            await self._ensure_sqlite()
            columns = ["id", "project_id", "filename", "file_size", "content_text", "status",
                       "category_edge", "measure_edge", "doc_type", "confidence",
                       "watts", "lumens", "tipo_equipo", "marca", "modelo",
                       "areas", "specialized_data", "uploaded_at"]
            if projection and "content_text" in projection and projection["content_text"] == 0:
                columns = [c for c in columns if c != "content_text"]
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT {', '.join(columns)} FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                if row:
                    return self._deserialize_file_row(dict(row))
            return None

    async def files_find(self, query: dict = None, projection: dict = None):
        if self.mode == "mongodb":
            proj = {k: 1 for k in projection.keys()} if projection else None
            cursor = db.files.find(query or {}, proj or {})
            return await cursor.to_list(500)
        else:
            await self._ensure_sqlite()
            columns = ["id", "project_id", "filename", "file_size", "status",
                       "category_edge", "measure_edge", "doc_type", "confidence",
                       "watts", "lumens", "tipo_equipo", "marca", "modelo",
                       "areas", "specialized_data", "uploaded_at"]
            if projection and "content_text" in projection and projection["content_text"] == 0:
                columns = [c for c in columns if c != "content_text"]
            sql = f"SELECT {', '.join(columns)} FROM files"
            params = []
            if query:
                where = [f"{k}=?" for k in query.keys()]
                params = list(query.values())
                sql += " WHERE " + " AND ".join(where)
            async with self.conn.execute(sql, params) as cur:
                rows = await cur.fetchall()
                return [self._deserialize_file_row(dict(r)) for r in rows]

    def _deserialize_file_row(self, row: dict) -> dict:
        if "areas" in row and row["areas"]:
            try:
                row["areas"] = json.loads(row["areas"])
            except:
                row["areas"] = None
        if "specialized_data" in row and row["specialized_data"]:
            try:
                row["specialized_data"] = json.loads(row["specialized_data"])
            except:
                row["specialized_data"] = None
        return row

    async def files_update_one(self, query: dict, update: dict):
        if self.mode == "mongodb":
            await db.files.update_one(query, update)
        else:
            await self._ensure_sqlite()
            set_dict = update.get("$set", {})
            if "areas" in set_dict and isinstance(set_dict["areas"], list):
                set_dict["areas"] = json.dumps(set_dict["areas"])
            if "specialized_data" in set_dict and isinstance(set_dict["specialized_data"], dict):
                set_dict["specialized_data"] = json.dumps(set_dict["specialized_data"])
            set_clause = ", ".join([f"{k}=?" for k in set_dict.keys()])
            params = list(set_dict.values())
            where = [f"{k}=?" for k in query.keys()]
            params.extend(list(query.values()))
            sql = f"UPDATE files SET {set_clause} WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()

    async def files_delete_one(self, query: dict):
        if self.mode == "mongodb":
            return await db.files.delete_one(query)
        else:
            await self._ensure_sqlite()
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                await self.conn.commit()
                return type('Result', (), {'deleted_count': cur.rowcount})()

    async def files_delete_many(self, query: dict):
        if self.mode == "mongodb":
            await db.files.delete_many(query)
        else:
            await self._ensure_sqlite()
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"DELETE FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params):
                await self.conn.commit()

    async def files_count_documents(self, query: dict):
        if self.mode == "mongodb":
            return await db.files.count_documents(query)
        else:
            await self._ensure_sqlite()
            where = [f"{k}=?" for k in query.keys()]
            params = list(query.values())
            sql = f"SELECT COUNT(*) as cnt FROM files WHERE {' AND '.join(where)}"
            async with self.conn.execute(sql, params) as cur:
                row = await cur.fetchone()
                return row["cnt"] if row else 0

class ProjectCreate(BaseModel):
    name: str
    typology: str

class ProjectResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    typology: str
    created_at: str
    file_count: int = 0
    processed_count: int = 0

class FileResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    project_id: str
    filename: str
    file_size: int
    status: str
    category_edge: Optional[str] = None
    measure_edge: Optional[str] = None
    doc_type: Optional[str] = None
    confidence: Optional[float] = None
    watts: Optional[float] = None
    lumens: Optional[float] = None
    tipo_equipo: Optional[str] = None
    marca: Optional[str] = None
    modelo: Optional[str] = None
    areas: Optional[list] = None
    specialized_data: Optional[dict] = None
    uploaded_at: str

# ── MOCK AI FUNCTIONS (Demo Mode) ───────────────────────────────────────

def classify_file_mock(content: str) -> dict:
    """Return deterministic mock classification based on filename/content patterns."""
    # Simple heuristic to make demo feel responsive
    content_lower = content.lower()
    if "lumin" in content_lower or "led" in content_lower or "lm" in content_lower:
        return {"category_edge": "ENERGY", "measure_edge": "EEM22", "doc_type": "ficha_tecnica", "confidence": 0.95}
    elif "hvac" in content_lower or "aire" in content_lower or " Split " in content or "VRF" in content:
        return {"category_edge": "ENERGY", "measure_edge": "EEM09", "doc_type": "ficha_tecnica", "confidence": 0.92}
    elif "plano" in content_lower or "floor" in content_lower or "architectural" in content_lower:
        return {"category_edge": "DESIGN", "measure_edge": "DESIGN", "doc_type": "plano", "confidence": 0.88}
    elif "griferia" in content_lower or "ducha" in content_lower or "inodoro" in content_lower or "sanitario" in content_lower:
        return {"category_edge": "WATER", "measure_edge": "WEM01", "doc_type": "ficha_tecnica", "confidence": 0.90}
    else:
        return {"category_edge": "MATERIALS", "measure_edge": "MRU01", "doc_type": "otro", "confidence": 0.75}

def extract_data_mock(content: str) -> dict:
    """Extract mock technical data from content."""
    content_lower = content.lower()
    watts = None
    lumens = None
    tipo_equipo = None
    marca = None
    modelo = None

    # Try to extract numbers using simple patterns
    import re
    watt_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:w|watts?|watt)', content_lower)
    if watt_match:
        watts = float(watt_match.group(1))

    lumen_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:lm|lumens?|lumen)', content_lower)
    if lumen_match:
        lumens = float(lumen_match.group(1))

    if "luminaria" in content_lower or "lampara" in content_lower or "led" in content_lower:
        tipo_equipo = "luminaria LED"
    elif "aire" in content_lower or "acondicionado" in content_lower or "split" in content_lower:
        tipo_equipo = "aire acondicionado"
    elif "grifo" in content_lower or "ducha" in content_lower:
        tipo_equipo = "griferia/ducha"

    # Common brands
    for brand in ["philips", "osram", "ge", "toshiba", "daikin", "carrier", "劲力"]:
        if brand in content_lower:
            marca = brand.capitalize()
            break

    return {
        "watts": watts,
        "lumens": lumens,
        "tipo_equipo": tipo_equipo,
        "marca": marca,
        "modelo": modelo
    }

def calculate_areas_mock(content: str) -> list:
    """Return mock area calculations for floor plans."""
    # Return some plausible demo areas
    return [
        {"nombre": "Oficina A", "area_m2": 45.5},
        {"nombre": "Sala de Reuniones", "area_m2": 22.0},
        {"nombre": "Pasillo", "area_m2": 12.3},
        {"nombre": "Baño", "area_m2": 6.5},
        {"nombre": "Cocina", "area_m2": 10.0},
    ]

# ── AI Processing Functions ─────────────────────────────────────────────

async def classify_file(content: str) -> dict:
    """Classify file using OpenAI GPT-4o or mock."""
    if not openai_client:
        return classify_file_mock(content)

    measures_list = ", ".join(EDGE_WBS.keys())
    prompt = f"""Clasifica este archivo tecnico de construccion.

Clasifica en:
1. Categoria EDGE (elige solo una): DESIGN, ENERGY, WATER, MATERIALS
2. Medida EDGE especifica (elige la mas probable de esta lista):
   {measures_list}
3. Tipo de documento: plano, ficha_tecnica, fotografia, memoria, factura, otro
4. Nivel de confianza (0 a 1)

Responde SOLO en JSON:
{{"category_edge": "...", "measure_edge": "...", "doc_type": "...", "confidence": 0.0}}

Contenido del archivo:
{content[:3000]}"""

    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Eres un consultor experto en certificacion EDGE. Responde SOLO en JSON valido. No expliques nada."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=500
        )
        result_text = response.choices[0].message.content.strip()

        if result_text.startswith("```"):
            result_text = result_text.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        return json.loads(result_text)
    except (json.JSONDecodeError, Exception) as e:
        logger.error(f"Failed to parse classification response: {e}")
        return classify_file_mock(content)


async def extract_data(content: str) -> dict:
    """Extract technical data using OpenAI GPT-4o or mock."""
    if not openai_client:
        return extract_data_mock(content)

    prompt = f"""Extrae la siguiente informacion si esta disponible:
- watts (numero)
- lumens (numero)
- tipo_equipo (ej: luminaria LED, aire acondicionado, bomba, etc.)
- marca
- modelo

Si no encuentras un dato, pon null. Convierte unidades a numeros (ej: "120 W" -> 120).

Responde SOLO en JSON:
{{"watts": null, "lumens": null, "tipo_equipo": null, "marca": null, "modelo": null}}

Contenido:
{content[:3000]}"""

    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Eres un ingeniero analizando fichas tecnicas de equipos. Responde SOLO en JSON valido."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=500
        )
        result_text = response.choices[0].message.content.strip()

        if result_text.startswith("```"):
            result_text = result_text.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        return json.loads(result_text)
    except (json.JSONDecodeError, Exception) as e:
        logger.error(f"Failed to parse extraction response: {e}")
        return extract_data_mock(content)


async def calculate_areas(content: str) -> list:
    """Calculate areas from floor plan text using OpenAI GPT-4o or mock."""
    if not openai_client:
        return calculate_areas_mock(content)

    prompt = f"""A partir del siguiente texto extraido de un plano (OCR), identifica espacios y sus dimensiones.
Calcula el area de cada espacio en m2.

Si hay largo y ancho, multiplica. Si no hay datos suficientes, ignora ese espacio.

Responde SOLO en JSON:
{{"espacios": [{{"nombre": "string", "area_m2": 0}}]}}

Texto del plano:
{content[:3000]}"""

    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Eres un arquitecto experto en interpretacion de planos. Responde SOLO en JSON valido."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=1000
        )
        result_text = response.choices[0].message.content.strip()

        if result_text.startswith("```"):
            result_text = result_text.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        data = json.loads(result_text)
        return data.get("espacios", [])
    except (json.JSONDecodeError, Exception) as e:
        logger.error(f"Failed to parse areas response: {e}")
        return calculate_areas_mock(content)


async def process_single_file_pipeline(file_doc: dict, job_id: str = None) -> dict:
    """Full processing pipeline for a single file."""
    content = file_doc.get("content_text", "")
    file_id = file_doc["id"]
    update = {}

    try:
        # Step 1: Classify
        classification = await classify_file(content)
        update = {
            "category_edge": classification.get("category_edge"),
            "measure_edge": classification.get("measure_edge"),
            "doc_type": classification.get("doc_type"),
            "confidence": classification.get("confidence"),
        }

        measure = classification.get("measure_edge", "")

        # Step 2: General data extraction
        extraction = await extract_data(content)
        update["watts"] = extraction.get("watts")
        update["lumens"] = extraction.get("lumens")
        update["tipo_equipo"] = extraction.get("tipo_equipo")
        update["marca"] = extraction.get("marca")
        update["modelo"] = extraction.get("modelo")

        # Step 3: Specialized processor based on measure
        # Pass API key only if available and not demo mode
        api_key = OPENAI_API_KEY if openai_client else None
        specialized = await run_specialized_processor(measure, content, api_key)
        if specialized:
            update["specialized_data"] = specialized

        # Step 4: Calculate areas for plans
        if classification.get("doc_type") == "plano":
            areas = await calculate_areas(content)
            update["areas"] = areas

        update["status"] = "processed"
        await udb.files_update_one({"id": file_id}, {"$set": update})
        return {"file_id": file_id, "filename": file_doc["filename"], "status": "processed", "measure": measure}

    except Exception as e:
        logger.error(f"Error processing file {file_doc['filename']}: {e}")
        await udb.files_update_one({"id": file_id}, {"$set": {"status": "error"}})
        return {"file_id": file_id, "filename": file_doc["filename"], "status": "error", "error": str(e)}


# --- API Routes ---

@api_router.get("/")
async def root():
    return {"message": "EDGE Document Processor API v2"}

# ═══════ EDGE Rules ═══════

@api_router.get("/edge-rules")
async def get_edge_rules():
    return get_all_rules()

@api_router.get("/edge-rules/{measure}")
async def get_edge_rule(measure: str):
    rule = get_rule(measure.upper())
    if not rule:
        raise HTTPException(status_code=404, detail=f"Medida {measure} no encontrada")
    return {measure.upper(): rule}

# ═══════ Projects ═══════

@api_router.post("/projects", response_model=ProjectResponse)
async def create_project(project: ProjectCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "name": project.name,
        "typology": project.typology,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "file_count": 0,
        "processed_count": 0,
    }
    await udb.projects_insert_one(doc)
    return ProjectResponse(**doc)

@api_router.get("/projects", response_model=List[ProjectResponse])
async def list_projects():
    projects = await udb.projects_find()
    result = []
    for p in projects:
        file_count = await udb.files_count_documents({"project_id": p["id"]})
        processed_count = await udb.files_count_documents({"project_id": p["id"], "status": "processed"})
        p["file_count"] = file_count
        p["processed_count"] = processed_count
        result.append(ProjectResponse(**p))
    return result

@api_router.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(project_id: str):
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    file_count = await udb.files_count_documents({"project_id": project_id})
    processed_count = await udb.files_count_documents({"project_id": project_id, "status": "processed"})
    project["file_count"] = file_count
    project["processed_count"] = processed_count
    return ProjectResponse(**project)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    result = await udb.projects_delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    await udb.files_delete_many({"project_id": project_id})
    # Cleanup jobs
    to_remove = [k for k, v in processing_jobs.items() if v.get("project_id") == project_id]
    for k in to_remove:
        del processing_jobs[k]
    return {"message": "Proyecto eliminado"}

# ═══════ Files ═══════

@api_router.post("/projects/{project_id}/files", response_model=FileResponse)
async def upload_file(project_id: str, file: UploadFile = File(...)):
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    content_bytes = await file.read()
    try:
        text_content = content_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text_content = content_bytes.decode("latin-1", errors="ignore")

    doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "filename": file.filename,
        "file_size": len(content_bytes),
        "content_text": text_content,
        "status": "pending",
        "category_edge": None,
        "measure_edge": None,
        "doc_type": None,
        "confidence": None,
        "watts": None,
        "lumens": None,
        "tipo_equipo": None,
        "marca": None,
        "modelo": None,
        "areas": None,
        "specialized_data": None,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await udb.files_insert_one(doc)
    # Return without content_text
    doc.pop("content_text", None)
    return FileResponse(**doc)

@api_router.get("/projects/{project_id}/files", response_model=List[FileResponse])
async def list_files(project_id: str):
    files = await udb.files_find({"project_id": project_id}, {"content_text": 0})
    return [FileResponse(**f) for f in files]

@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str):
    result = await udb.files_delete_one({"id": file_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return {"message": "Archivo eliminado"}

# ═══════ AI Processing with Batch Progress ═══════

@api_router.post("/projects/{project_id}/process-edge")
async def process_edge_project(project_id: str):
    """Full EDGE automation: classify + extract + specialized analysis + validate."""
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    # Get ALL files (pending + already processed to reprocess)
    files = await udb.files_find({"project_id": project_id})

    if not files:
        raise HTTPException(status_code=400, detail="No hay archivos en el proyecto")

    # Create job tracker
    job_id = str(uuid.uuid4())
    processing_jobs[job_id] = {
        "project_id": project_id,
        "status": "running",
        "total": len(files),
        "processed": 0,
        "current_file": "",
        "current_step": "Iniciando...",
        "results": [],
        "started_at": datetime.now(timezone.utc).isoformat(),
    }

    # Start background processing
    asyncio.create_task(_run_batch_processing(job_id, files))

    return {"job_id": job_id, "total_files": len(files), "status": "running"}


async def _run_batch_processing(job_id: str, files: list):
    """Background task for batch processing."""
    job = processing_jobs[job_id]
    try:
        for i, f in enumerate(files):
            job["current_file"] = f["filename"]
            job["current_step"] = f"Clasificando ({i+1}/{len(files)})"
            job["processed"] = i

            result = await process_single_file_pipeline(f, job_id)
            job["results"].append(result)

        job["status"] = "completed"
        job["processed"] = len(files)
        job["current_step"] = "Completado"
        job["current_file"] = ""
        job["completed_at"] = datetime.now(timezone.utc).isoformat()
    except Exception as e:
        logger.error(f"Batch processing error: {e}")
        job["status"] = "error"
        job["current_step"] = f"Error: {str(e)}"


@api_router.get("/projects/{project_id}/process-status/{job_id}")
async def get_process_status(project_id: str, job_id: str):
    """Get real-time batch processing status."""
    job = processing_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return {
        "job_id": job_id,
        "status": job["status"],
        "total": job["total"],
        "processed": job["processed"],
        "percent": round((job["processed"] / job["total"]) * 100) if job["total"] > 0 else 0,
        "current_file": job["current_file"],
        "current_step": job["current_step"],
        "results": job.get("results", []),
    }

# Legacy simple process (backward compat)
@api_router.post("/projects/{project_id}/process")
async def process_project_files(project_id: str):
    files = await udb.files_find({"project_id": project_id, "status": "pending"})

    if not files:
        raise HTTPException(status_code=400, detail="No hay archivos pendientes de procesar")

    results = []
    for f in files:
        result = await process_single_file_pipeline(f)
        results.append(result)

    return {"processed": len(results), "results": results}

@api_router.post("/files/{file_id}/process")
async def process_single_file(file_id: str):
    f = await udb.files_find_one({"id": file_id})
    if not f:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    result = await process_single_file_pipeline(f)
    return result

# ═══════ WBS Validation (Deterministic) ═══════

@api_router.get("/projects/{project_id}/wbs-validation")
async def get_wbs_validation(project_id: str):
    """Deterministic WBS validation against EDGE rules."""
    files = await udb.files_find(
        {"project_id": project_id, "status": "processed"},
        {"content_text": 0}
    )

    validation = validate_project_wbs(files)
    coverage = get_project_coverage(files)

    return {
        "measures": validation,
        "coverage": coverage,
        "total_files": await udb.files_count_documents({"project_id": project_id}),
        "processed_files": len(files),
    }

# ═══════ KPIs ═══════

@api_router.get("/projects/{project_id}/kpis")
async def get_project_kpis(project_id: str):
    """Get KPIs per measure for the project."""
    files = await udb.files_find(
        {"project_id": project_id, "status": "processed"},
        {"content_text": 0}
    )

    kpis = {}

    # EEM22/EEM23 - Lighting efficacy
    lighting_files = [f for f in files if f.get("measure_edge") in ("EEM22", "EEM23") and f.get("specialized_data")]
    if lighting_files:
        total_lm = 0
        total_w = 0
        all_luminaires = []
        alertas = []
        for f in lighting_files:
            sd = f["specialized_data"]
            total_lm += sd.get("total_lumens", 0)
            total_w += sd.get("total_watts", 0)
            all_luminaires.extend(sd.get("luminarias", []))
            alertas.extend(sd.get("alertas", []))

        eficacia = round(total_lm / total_w, 2) if total_w > 0 else 0
        kpis["EEM22"] = {
            "nombre": "Eficacia Luminosa Global",
            "valor": eficacia,
            "unidad": "lm/W",
            "umbral_edge": 90,
            "cumple": eficacia >= 90,
            "total_lumens": total_lm,
            "total_watts": total_w,
            "num_luminarias": len(all_luminaires),
            "alertas": list(set(alertas)),
        }

    # EEM09 - HVAC COP
    hvac_files = [f for f in files if f.get("measure_edge") == "EEM09" and f.get("specialized_data")]
    if hvac_files:
        cops = []
        for f in hvac_files:
            sd = f["specialized_data"]
            for eq in sd.get("equipos", []):
                if eq.get("cop"):
                    cops.append(eq["cop"])
        avg_cop = round(sum(cops) / len(cops), 2) if cops else 0
        kpis["EEM09"] = {
            "nombre": "COP Promedio HVAC",
            "valor": avg_cop,
            "unidad": "COP",
            "umbral_edge": 3.0,
            "cumple": avg_cop >= 3.0,
            "num_equipos": sum(len(f["specialized_data"].get("equipos", [])) for f in hvac_files),
        }

    # Water fixtures
    water_files = [f for f in files if f.get("measure_edge") in ("WEM01", "WEM02") and f.get("specialized_data")]
    if water_files:
        for f in water_files:
            measure = f["measure_edge"]
            sd = f["specialized_data"]
            if measure not in kpis:
                kpis[measure] = {
                    "nombre": "Flujo Promedio" if measure == "WEM01" else "Descarga Promedio",
                    "valor": sd.get("flujo_promedio", 0),
                    "unidad": "lpm" if measure == "WEM01" else "lpd",
                    "num_aparatos": len(sd.get("aparatos", [])),
                }

    return kpis

# ═══════ Extracted Data ═══════

@api_router.get("/projects/{project_id}/extracted-data")
async def get_extracted_data(project_id: str):
    files = await udb.files_find(
        {"project_id": project_id, "status": "processed"},
        {"content_text": 0}
    )
    return files

# ═══════ EDGE Status (enhanced) ═══════

@api_router.get("/projects/{project_id}/edge-status")
async def get_edge_status(project_id: str):
    files = await udb.files_find(
        {"project_id": project_id, "status": "processed"},
        {"content_text": 0}
    )

    categories = {}
    measures = {}
    for f in files:
        cat = f.get("category_edge", "UNKNOWN")
        meas = f.get("measure_edge", "UNKNOWN")
        categories[cat] = categories.get(cat, 0) + 1
        measures[meas] = measures.get(meas, 0) + 1

    # Use deterministic WBS validation instead of AI
    validation = validate_project_wbs(files)
    faltantes = [
        {"medida": k, "faltan": v["faltantes"]}
        for k, v in validation.items()
        if v["estado"] == "incompleto"
    ]

    total = await udb.files_count_documents({"project_id": project_id})
    processed = await udb.files_count_documents({"project_id": project_id, "status": "processed"})

    return {
        "categories": categories,
        "measures": measures,
        "faltantes": faltantes,
        "total_files": total,
        "processed_files": processed,
        "wbs_validation": validation,
    }

# ═══════ Export Enhanced ═══════

@api_router.get("/projects/{project_id}/export-excel")
async def export_excel(project_id: str):
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    files = await udb.files_find({"project_id": project_id}, {"content_text": 0})

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Sheet 1: Classification
    ws = wb.active
    ws.title = "Clasificacion EDGE"
    headers = ["Archivo", "Categoria", "Medida", "Tipo Doc", "Confianza", "Watts", "Lumens", "Equipo", "Marca", "Modelo", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(files, 2):
        vals = [
            f.get("filename", ""), f.get("category_edge", ""), f.get("measure_edge", ""),
            f.get("doc_type", ""), f.get("confidence", ""), f.get("watts", ""),
            f.get("lumens", ""), f.get("tipo_equipo", ""), f.get("marca", ""),
            f.get("modelo", ""), f.get("status", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # Sheet 2: WBS Validation
    processed_files = [f for f in files if f.get("status") == "processed"]
    validation = validate_project_wbs(processed_files)
    if validation:
        ws2 = wb.create_sheet("Validacion WBS")
        headers2 = ["Medida", "Categoria", "Nombre", "Estado", "Progreso", "Docs Requeridos", "Docs Disponibles", "Faltantes"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row, (measure, data) in enumerate(sorted(validation.items()), 2):
            ws2.cell(row=row, column=1, value=measure).border = thin_border
            ws2.cell(row=row, column=2, value=data.get("categoria", "")).border = thin_border
            ws2.cell(row=row, column=3, value=data.get("nombre", "")).border = thin_border
            ws2.cell(row=row, column=4, value=data.get("estado", "")).border = thin_border
            ws2.cell(row=row, column=5, value=f"{int(data.get('progreso', 0)*100)}%").border = thin_border
            ws2.cell(row=row, column=6, value=", ".join(data.get("documentos_requeridos", []))).border = thin_border
            ws2.cell(row=row, column=7, value=", ".join(data.get("documentos_disponibles", []))).border = thin_border
            ws2.cell(row=row, column=8, value=", ".join(data.get("faltantes", []))).border = thin_border
        for col in range(1, len(headers2) + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Sheet 3: EEM22 Luminaires
    eem22_files = [f for f in files if f.get("measure_edge") in ("EEM22", "EEM23") and f.get("specialized_data")]
    if eem22_files:
        ws3 = wb.create_sheet("EEM22 Luminarias")
        headers3 = ["Archivo", "ID", "Modelo", "Cantidad", "Lumens", "Watts", "Eficiencia (lm/W)", "Notas"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row = 2
        for f in eem22_files:
            for lum in f["specialized_data"].get("luminarias", []):
                ws3.cell(row=row, column=1, value=f["filename"]).border = thin_border
                ws3.cell(row=row, column=2, value=lum.get("id", "")).border = thin_border
                ws3.cell(row=row, column=3, value=lum.get("modelo", "")).border = thin_border
                ws3.cell(row=row, column=4, value=lum.get("cantidad", 0)).border = thin_border
                ws3.cell(row=row, column=5, value=lum.get("lumens", 0)).border = thin_border
                ws3.cell(row=row, column=6, value=lum.get("watts", 0)).border = thin_border
                ws3.cell(row=row, column=7, value=lum.get("eficiencia", 0)).border = thin_border
                ws3.cell(row=row, column=8, value=lum.get("notas", "")).border = thin_border
                row += 1
            # Add summary row
            sd = f["specialized_data"]
            ws3.cell(row=row, column=1, value="").border = thin_border
            ws3.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
            ws3.cell(row=row, column=4, value=sd.get("total_luminarias", 0)).font = Font(bold=True)
            ws3.cell(row=row, column=5, value=sd.get("total_lumens", 0)).font = Font(bold=True)
            ws3.cell(row=row, column=6, value=sd.get("total_watts", 0)).font = Font(bold=True)
            eficacia = sd.get("eficacia_global", 0)
            ws3.cell(row=row, column=7, value=eficacia).font = Font(bold=True, color="00AA00" if eficacia >= 90 else "FF0000")
            row += 1
        for col in range(1, len(headers3) + 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # Sheet 4: Areas
    files_with_areas = [f for f in files if f.get("areas")]
    if files_with_areas:
        ws4 = wb.create_sheet("Areas")
        headers4 = ["Archivo", "Espacio", "Area m2"]
        for col, h in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row = 2
        for f in files_with_areas:
            for area in f["areas"]:
                ws4.cell(row=row, column=1, value=f["filename"]).border = thin_border
                ws4.cell(row=row, column=2, value=area.get("nombre", "")).border = thin_border
                ws4.cell(row=row, column=3, value=area.get("area_m2", "")).border = thin_border
                row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{project['name'].replace(' ', '_')}_EDGE.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ═══════ PDF Report ═══════

@api_router.get("/projects/{project_id}/export-pdf")
async def export_pdf(project_id: str):
    """Generate professional EDGE certification PDF report."""
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    files = await udb.files_find({"project_id": project_id}, {"content_text": 0})

    processed_files = [f for f in files if f.get("status") == "processed"]

    # Get validation and KPIs data
    validation_data = validate_project_wbs(processed_files)
    coverage_data = get_project_coverage(processed_files)

    validation = {
        "measures": validation_data,
        "coverage": coverage_data,
        "total_files": len(files),
        "processed_files": len(processed_files),
    }

    # Build KPIs
    kpis = {}
    lighting_files = [f for f in processed_files if f.get("measure_edge") in ("EEM22", "EEM23") and f.get("specialized_data")]
    if lighting_files:
        total_lm = sum(f["specialized_data"].get("total_lumens", 0) for f in lighting_files)
        total_w = sum(f["specialized_data"].get("total_watts", 0) for f in lighting_files)
        eficacia = round(total_lm / total_w, 2) if total_w > 0 else 0
        all_alerts = []
        for f in lighting_files:
            all_alerts.extend(f["specialized_data"].get("alertas", []))
        kpis["EEM22"] = {
            "nombre": "Eficacia Luminosa Global",
            "valor": eficacia,
            "unidad": "lm/W",
            "umbral_edge": 90,
            "cumple": eficacia >= 90,
            "alertas": list(set(all_alerts)),
        }

    hvac_files = [f for f in processed_files if f.get("measure_edge") == "EEM09" and f.get("specialized_data")]
    if hvac_files:
        cops = []
        for f in hvac_files:
            for eq in f["specialized_data"].get("equipos", []):
                if eq.get("cop"):
                    cops.append(eq["cop"])
        avg_cop = round(sum(cops) / len(cops), 2) if cops else 0
        kpis["EEM09"] = {
            "nombre": "COP Promedio HVAC",
            "valor": avg_cop,
            "unidad": "COP",
            "umbral_edge": 3.0,
            "cumple": avg_cop >= 3.0,
            "alertas": [],
        }

    water_files = [f for f in processed_files if f.get("measure_edge") in ("WEM01", "WEM02") and f.get("specialized_data")]
    for f in water_files:
        measure = f["measure_edge"]
        sd = f["specialized_data"]
        if measure not in kpis:
            kpis[measure] = {
                "nombre": "Flujo Promedio" if measure == "WEM01" else "Descarga Promedio",
                "valor": sd.get("flujo_promedio", 0),
                "unidad": "lpm" if measure == "WEM01" else "lpd",
                "alertas": [],
            }

    buffer = generate_edge_report(project, files, validation, kpis)

    filename = f"{project['name'].replace(' ', '_')}_Reporte_EDGE.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db():
    if MONGO_URL:
        client.close()
        logger.info("MongoDB connection closed")
    elif udb.conn:
        await udb.conn.close()
        logger.info("SQLite connection closed")
