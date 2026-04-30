from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.db.database import udb
from app.services.edge_rules import validate_project_wbs, get_project_coverage
from app.services.pdf_generator import generate_edge_report

router = APIRouter()

@router.get("/projects/{project_id}/export-excel")
async def export_excel(project_id: str):
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    files = await udb.files_find({"project_id": project_id}, {"content_text": 0})

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    ws = wb.active
    ws.title = "Clasificacion EDGE"
    headers = ["Archivo", "Categoria", "Medida", "Tipo Doc", "Confianza", "Watts", "Lumens", "Equipo", "Marca", "Modelo", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(files, 2):
        vals = [
            f.get("filename", ""), f.get("category_edge", ""), f.get("measure_edge", ""),
            f.get("doc_type", ""), f.get("confidence", ""), f.get("watts", ""),
            f.get("lumens", ""), f.get("tipo_equipo", ""), f.get("marca", ""),
            f.get("modelo", ""), f.get("status", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    processed_files = [f for f in files if f.get("status") == "processed"]
    validation = validate_project_wbs(processed_files)
    if validation:
        ws2 = wb.create_sheet("Validacion WBS")
        headers2 = ["Medida", "Categoria", "Nombre", "Estado", "Progreso", "Docs Requeridos", "Docs Disponibles", "Faltantes"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row, (measure, data) in enumerate(sorted(validation.items()), 2):
            ws2.cell(row=row, column=1, value=measure).border = thin_border
            ws2.cell(row=row, column=2, value=data.get("categoria", "")).border = thin_border
            ws2.cell(row=row, column=3, value=data.get("nombre", "")).border = thin_border
            ws2.cell(row=row, column=4, value=data.get("estado", "")).border = thin_border
            ws2.cell(row=row, column=5, value=f"{int(data.get('progreso', 0)*100)}%").border = thin_border
            ws2.cell(row=row, column=6, value=", ".join(data.get("documentos_requeridos", []))).border = thin_border
            ws2.cell(row=row, column=7, value=", ".join(data.get("documentos_disponibles", []))).border = thin_border
            ws2.cell(row=row, column=8, value=", ".join(data.get("faltantes", []))).border = thin_border
        for col in range(1, len(headers2) + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    eem22_files = [f for f in files if f.get("measure_edge") in ("EEM22", "EEM23") and f.get("specialized_data")]
    if eem22_files:
        ws3 = wb.create_sheet("EEM22 Luminarias")
        headers3 = ["Archivo", "ID", "Modelo", "Cantidad", "Lumens", "Watts", "Eficiencia (lm/W)", "Notas"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row = 2
        for f in eem22_files:
            for lum in f["specialized_data"].get("luminarias", []):
                ws3.cell(row=row, column=1, value=f["filename"]).border = thin_border
                ws3.cell(row=row, column=2, value=lum.get("id", "")).border = thin_border
                ws3.cell(row=row, column=3, value=lum.get("modelo", "")).border = thin_border
                ws3.cell(row=row, column=4, value=lum.get("cantidad", 0)).border = thin_border
                ws3.cell(row=row, column=5, value=lum.get("lumens", 0)).border = thin_border
                ws3.cell(row=row, column=6, value=lum.get("watts", 0)).border = thin_border
                ws3.cell(row=row, column=7, value=lum.get("eficiencia", 0)).border = thin_border
                ws3.cell(row=row, column=8, value=lum.get("notas", "")).border = thin_border
                row += 1
            sd = f["specialized_data"]
            ws3.cell(row=row, column=1, value="").border = thin_border
            ws3.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
            ws3.cell(row=row, column=4, value=sd.get("total_luminarias", 0)).font = Font(bold=True)
            ws3.cell(row=row, column=5, value=sd.get("total_lumens", 0)).font = Font(bold=True)
            ws3.cell(row=row, column=6, value=sd.get("total_watts", 0)).font = Font(bold=True)
            eficacia = sd.get("eficacia_global", 0)
            ws3.cell(row=row, column=7, value=eficacia).font = Font(bold=True, color="00AA00" if eficacia >= 90 else "FF0000")
            row += 1
        for col in range(1, len(headers3) + 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    files_with_areas = [f for f in files if f.get("areas")]
    if files_with_areas:
        ws4 = wb.create_sheet("Areas")
        headers4 = ["Archivo", "Espacio", "Area m2"]
        for col, h in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row = 2
        for f in files_with_areas:
            for area in f["areas"]:
                ws4.cell(row=row, column=1, value=f["filename"]).border = thin_border
                ws4.cell(row=row, column=2, value=area.get("nombre", "")).border = thin_border
                ws4.cell(row=row, column=3, value=area.get("area_m2", "")).border = thin_border
                row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{project['name'].replace(' ', '_')}_EDGE.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/projects/{project_id}/export-pdf")
async def export_pdf(project_id: str):
    project = await udb.projects_find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    files = await udb.files_find({"project_id": project_id}, {"content_text": 0})
    processed_files = [f for f in files if f.get("status") == "processed"]

    validation_data = validate_project_wbs(processed_files)
    coverage_data = get_project_coverage(processed_files)

    validation = {
        "measures": validation_data,
        "coverage": coverage_data,
        "total_files": len(files),
        "processed_files": len(processed_files),
    }

    kpis = {}
    lighting_files = [f for f in processed_files if f.get("measure_edge") in ("EEM22", "EEM23") and f.get("specialized_data")]
    if lighting_files:
        total_lm = sum(f["specialized_data"].get("total_lumens", 0) for f in lighting_files)
        total_w = sum(f["specialized_data"].get("total_watts", 0) for f in lighting_files)
        eficacia = round(total_lm / total_w, 2) if total_w > 0 else 0
        all_alerts = []
        for f in lighting_files:
            all_alerts.extend(f["specialized_data"].get("alertas", []))
        kpis["EEM22"] = {
            "nombre": "Eficacia Luminosa Global",
            "valor": eficacia,
            "unidad": "lm/W",
            "umbral_edge": 90,
            "cumple": eficacia >= 90,
            "alertas": list(set(all_alerts)),
        }

    hvac_files = [f for f in processed_files if f.get("measure_edge") == "EEM09" and f.get("specialized_data")]
    if hvac_files:
        cops = []
        for f in hvac_files:
            for eq in f["specialized_data"].get("equipos", []):
                if eq.get("cop"):
                    cops.append(eq["cop"])
        avg_cop = round(sum(cops) / len(cops), 2) if cops else 0
        kpis["EEM09"] = {
            "nombre": "COP Promedio HVAC",
            "valor": avg_cop,
            "unidad": "COP",
            "umbral_edge": 3.0,
            "cumple": avg_cop >= 3.0,
            "alertas": [],
        }

    water_files = [f for f in processed_files if f.get("measure_edge") in ("WEM01", "WEM02") and f.get("specialized_data")]
    for f in water_files:
        measure = f["measure_edge"]
        sd = f["specialized_data"]
        if measure not in kpis:
            kpis[measure] = {
                "nombre": "Flujo Promedio" if measure == "WEM01" else "Descarga Promedio",
                "valor": sd.get("flujo_promedio", 0),
                "unidad": "lpm" if measure == "WEM01" else "lpd",
                "alertas": [],
            }

    buffer = generate_edge_report(project, files, validation, kpis)

    filename = f"{project['name'].replace(' ', '_')}_Reporte_EDGE.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
